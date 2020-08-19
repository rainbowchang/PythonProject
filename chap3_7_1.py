# # from kkb_tools import open_file
# import csv
#
# # 需要写入的数据
# score1 = ['math', 95]
# score2 = ['english', 90]
#
# # 打开文件，追加a, newline=""，可以删掉行与行之间的空格
# file= open("C:/Users/Administrator/Desktop/score.csv", "a", newline="")
#
# # 设定写入模式
# csv_write = csv.writer(file)
#
# # 写入具体内容
# csv_write.writerow(score1)
# csv_write.writerow(score2)
# file.close()
# # open_file('score.csv')



# import openpyxl
# # 引用openpyxl
# wb = openpyxl.Workbook()
# # 利用openpyxl.Workbook()函数创建新的workbook（工作薄）对象，就是创建新的空的Excel文件。
# sheet = wb.active
# # wb.active就是获取这个工作薄的活动表，通常就是第一个工作簿，也就是我们在上面的图片中看到的sheet1。
# sheet.title = 'kaikeba'
# # 可以用.title给工作表重命名。现在第一个工作表的名称就会由原来默认的“sheet1”改为"kaikeba"。
# sheet['A1'] = 'kaikeba'
# # 向单个单元格写入数据
# score1 = ['math', 96]
# sheet.append(score1)
# # 写入整行的数据，变量类型是一个列表
# wb.save('C:/Users/Administrator/Desktop/score.xlsx')
# # 保存修改的Excel
# wb.close()
# # 关闭Excel



# import openpyxl
# wb = openpyxl.load_workbook('C:/Users/Administrator/Desktop/score.xlsx')
# # 打开的指定的工作簿
# sheet = wb['kaikeba']
# # 指定读取的工作表的名称
# A1_value = sheet['b2'].value
# print(A1_value)
# # 获取


import requests
import openpyxl

url = 'https://c.y.qq.com/soso/fcgi-bin/client_search_cp'
wb=openpyxl.Workbook()
#创建工作薄
sheet=wb.active
#获取工作薄的活动表
sheet.title='songs_mayday'
#工作表重命名
column_name = ['歌曲名','所属专辑','播放时长','播放链接']
sheet.append(column_name)
for x in range(0,3):
    params = {
        'ct': '24',
        'qqmusic_ver': '1298',
        'new_json': '1',
        'remoteplace': 'sizer.yqq.song_next',
        'searchid': '64405487069162918',
        't': '0',
        'aggr': '1',
        'cr': '1',
        'catZhida': '1',
        'lossless': '0',
        'flag_qc': '0',
        'p': str(x + 1),
        'n': '20',
        'w': '五月天',
        'g_tk': '5381',
        'loginUin': '0',
        'hostUin': '0',
        'format': 'json',
        'inCharset': 'utf8',
        'outCharset': 'utf-8',
        'notice': '0',
        'platform': 'yqq.json',
        'needNewCode': '0'
    }
    # 将参数封装为字典
    res_music = requests.get(url, params=params)
    # 调用get方法，下载这个列表
    json_music = res_music.json()
    # 使用json()方法，将response对象，转为列表/字典
    list_music = json_music['data']['song']['list']
    # 一层一层地取字典，获取歌单列表
    for music in list_music:
        name = music['name']
        # 以name为键，查找歌曲名，把歌曲名赋值给name
        album = music['album']['name']
        # 查找专辑名，把专辑名赋给album
        time = music['interval']
        # 查找播放时长，把时长赋值给time
        link = 'https://y.qq.com/n/yqq/song/' + str(music['file']['media_mid']) + '.html\n\n'
        # 查找播放链接，把链接赋值给link
        sheet.append([name, album, time, link])
        # 把name、album、time和link写成列表，用append函数多行写入Excel

wb.save('C:/Users/Administrator/Desktop/mayday.xlsx')
wb.close()
#最后保存并关闭这个Excel文件





